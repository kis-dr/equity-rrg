# -*- coding: utf-8 -*-
"""
build.py
--------
Equity_RRG.xlsx -> docs/data.json 변환 스크립트.

엑셀 'RRG' 시트 구조 (실측 기준):
  - 1행: 블록 제목  ('상대강도' / '1YR EXCESS RETURN' , '모멘텀')
  - 2행: 헤더
        · REL_RET 블록: A열='REL_RET', B~N열(2~14)=13개 시리즈명
        · RSI     블록: P열='RSI',     Q~AC열(17~29)=동일 13개 시리즈명
  - 3~30행: 실제 데이터 (28주, 주간). 31행 이하는 스크래치라 무시.
        · REL_RET: A열=날짜, B~N열=1년 초과수익률(%)   -> Y축(상대강도)
        · RSI    : P열=날짜, Q~AC열=RSI(0~100)          -> X축(모멘텀)

축 정의(사용자 확정): X = RSI, Y = 1년 초과수익률.
스케일: 원본 유지. 기준선 X=50, Y=0.

의존성: openpyxl  (pip install openpyxl)
실행: python build.py
"""

import json
import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

# ---- 경로 (레포 루트 기준 상대경로) -----------------------------------------
ROOT = Path(__file__).resolve().parent
SRC_XLSX = ROOT / "data" / "Equity_RRG.xlsx"
OUT_JSON = ROOT / "docs" / "data.json"

SHEET = "RRG"

# 데이터 영역 (1-indexed, 엑셀 좌표 기준)
DATA_FIRST_ROW = 3
DATA_LAST_ROW = 30          # 31행부터는 잔여 스크래치라 제외
HEADER_ROW = 2

# REL_RET 블록: 날짜 A(1), 시리즈 B~N(2~14)
REL_DATE_COL = 1
REL_FIRST_COL = 2
REL_LAST_COL = 14
# RSI 블록: 날짜 P(16), 시리즈 Q~AC(17~29)
RSI_DATE_COL = 16
RSI_FIRST_COL = 17
RSI_LAST_COL = 29


def _cell(ws, r, c):
    return ws.cell(row=r, column=c).value


def main():
    if not SRC_XLSX.exists():
        raise FileNotFoundError(f"소스 엑셀 없음: {SRC_XLSX}")

    wb = load_workbook(SRC_XLSX, data_only=True)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"'{SHEET}' 시트 없음. 존재 시트: {wb.sheetnames}")
    ws = wb[SHEET]

    # --- 시리즈명 (두 블록이 동일해야 정상) ---
    rel_names = [_cell(ws, HEADER_ROW, c) for c in range(REL_FIRST_COL, REL_LAST_COL + 1)]
    rsi_names = [_cell(ws, HEADER_ROW, c) for c in range(RSI_FIRST_COL, RSI_LAST_COL + 1)]
    rel_names = [str(n).strip() for n in rel_names]
    rsi_names = [str(n).strip() for n in rsi_names]
    if rel_names != rsi_names:
        print("[경고] REL_RET / RSI 블록 시리즈명이 불일치합니다. RSI 블록 순서를 기준으로 정렬합니다.")
        print("  REL_RET:", rel_names)
        print("  RSI    :", rsi_names)

    names = rel_names  # 13개

    # --- 날짜 (REL 블록 A열 기준) 및 두 블록 날짜 일치 검증 ---
    dates = []
    for r in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        dv = _cell(ws, r, REL_DATE_COL)
        if isinstance(dv, dt.datetime):
            dates.append(dv.date())
        elif isinstance(dv, dt.date):
            dates.append(dv)
        else:
            raise ValueError(f"{r}행 A열 날짜 아님: {dv!r}")

    # --- 값 읽기: series[name] = { 'rsi':[...], 'exret':[...] } (날짜 정렬) ---
    series = []
    x_all, y_all = [], []
    for i, name in enumerate(names):
        rel_col = REL_FIRST_COL + i
        rsi_col = RSI_FIRST_COL + i
        points = []  # [x(rsi), y(exret)] per date
        for k, r in enumerate(range(DATA_FIRST_ROW, DATA_LAST_ROW + 1)):
            y = _cell(ws, r, rel_col)   # 초과수익률 -> Y
            x = _cell(ws, r, rsi_col)   # RSI        -> X
            if x is None or y is None:
                # 결측이면 해당 주는 null 로 (프론트에서 끊어서 처리)
                points.append(None)
                continue
            x = float(x)
            y = float(y)
            points.append([round(x, 4), round(y, 4)])
            x_all.append(x)
            y_all.append(y)
        series.append({"name": name, "points": points})

    # --- 축 범위 (원본 스케일 유지, 기준선이 보이도록 패딩) ---
    def padded_range(vals, center, pad_ratio=0.08):
        lo, hi = min(vals), max(vals)
        # 기준선(center)이 항상 범위 안에 들어오도록 보정
        lo = min(lo, center)
        hi = max(hi, center)
        span = hi - lo
        pad = span * pad_ratio if span > 0 else 1.0
        return round(lo - pad, 2), round(hi + pad, 2)

    x_min, x_max = padded_range(x_all, 50.0)
    y_min, y_max = padded_range(y_all, 0.0)

    out = {
        "meta": {
            "generated": dt.datetime.now().isoformat(timespec="seconds"),
            "source": SRC_XLSX.name,
            "x_metric": "RSI (모멘텀)",
            "y_metric": "1년 초과수익률 % (상대강도)",
            "x_center": 50.0,
            "y_center": 0.0,
            "x_min": x_min, "x_max": x_max,
            "y_min": y_min, "y_max": y_max,
            "n_weeks": len(dates),
            "n_series": len(names),
        },
        "dates": [d.isoformat() for d in dates],
        "series": series,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[완료] {OUT_JSON}  |  {len(names)}개 시리즈 x {len(dates)}주")
    print(f"       X(RSI) 범위 {x_min}~{x_max}  /  Y(초과수익률) 범위 {y_min}~{y_max}")


if __name__ == "__main__":
    main()
